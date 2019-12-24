from maediprojects.lib import util
from openpyxl.styles import Color, PatternFill, Font, Border, Protection, Alignment
from openpyxl.styles.borders import Border, Side
from formatting import yellowFill, orangeFill

def cover_sheet(_self):
    ws = _self.wb.create_sheet(u"Instructions")
    ws["A1"] = u"Instructions"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = u"Thank you for providing your data to AMCU! Capturing high quality data is vitally important to strengthening aid effectiveness in Liberia."
    ws["A2"].font = Font(size=14)
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:K4")
    ws["A6"] = u"Currency"
    ws["A6"].font = Font(bold=True, size=14)
    ws["A7"] = u"Please provide your data in the currency stated above. Please contact AMCU if you would like this template in a different currency (your existing data will be exported in your desired currency to ensure consistency)."
    ws["A7"].font = Font(size=14)
    ws["A7"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A7:K10")
    ws["C6"] = _self.template_currency
    ws["C6"].font = Font(size=14)
    yellow_fill = PatternFill(start_color='FFFF00',
                         end_color='FFFF00',
                         fill_type = 'solid')
    thin_border = Border(left=Side(style='thin'),
                 right=Side(style='thin'),
                 top=Side(style='thin'),
                 bottom=Side(style='thin'))

    ws["C6"].fill = yellow_fill
    ws["C6"].border = thin_border
    ws["A12"] = u"How to fill in this template"
    ws["A12"].font = Font(bold=True, size=14)
    ws["A13"] = u"On the next sheet, you will see a list of projects currently known to AMCU. Fill out the sheet as follows (if you have any further questions, contact AMCU):"
    ws["A13"].font = Font(size=14)
    ws["A13"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A13:K15")
    return ws


def mtef(_self, ws):
    ws["A16"] = u"Counterpart funding"
    ws["A16"].font = Font(size=14)
    ws["A16"].alignment = Alignment(vertical="top")
    ws["A16"].fill = orangeFill
    ws.row_dimensions[16].height=22
    ws.merge_cells("A16:C17")
    ws["D16"] = u"The amount that GoL has agreed to provide as counterpart funding for this project, for the coming Fiscal Year, in USD."
    ws["D16"].font = Font(size=14)
    ws["D16"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("D16:K17")

    ws["A18"] = u"MTEF Projections"
    ws["A18"].font = Font(size=14)
    ws["A18"].fill = yellowFill
    ws.merge_cells("A18:C18")
    ws["D18"] = u"The amount you plan to spend in each of the next 3 Fiscal Years, in {}.".format(_self.template_currency)
    ws["D18"].font = Font(size=14)
    ws.merge_cells("D18:K18")

    ws["A19"] = u"Other project data"
    ws["A19"].font = Font(size=14)
    ws.merge_cells("A19:C19")
    ws["D19"] = u"Please check other project data and update it as required."
    ws["D19"].font = Font(size=14)
    ws.merge_cells("D19:K19")

    ws["A20"] = u"New projects"
    ws["A20"].font = Font(size=14)
    ws["A20"].alignment = Alignment(vertical="top")
    ws.merge_cells("A20:C21")
    ws["D20"] = u"For new projects, add new rows at the bottom of the sheet. Leave the ID column blank."
    ws["D20"].font = Font(size=14)
    ws["D20"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[20].height=22
    ws.merge_cells("D20:K21")
    return ws


def disbursements(_self, ws):
    ws["A16"] = util.previous_fy_fq()
    ws["A16"].font = Font(size=14)
    ws["A16"].fill = yellowFill
    ws.merge_cells("A16:C16")
    ws["D16"] = u"The amount you spent on this project in the last fiscal quarter, in {}.".format(_self.template_currency)
    ws["D16"].font = Font(size=14)
    ws.merge_cells("D16:K16")

    ws["A17"] = u"Other project data"
    ws["A17"].font = Font(size=14)
    ws.merge_cells("A17:C17")
    ws["D17"] = u"Please check other project data and update it as required."
    ws["D17"].font = Font(size=14)
    ws.merge_cells("D17:K17")

    ws["A18"] = u"New projects"
    ws["A18"].font = Font(size=14)
    ws["A18"].alignment = Alignment(vertical="top")
    ws.merge_cells("A18:C19")
    ws["D18"] = u"For new projects, add new rows at the bottom of the sheet. Leave the ID column blank."
    ws["D18"].font = Font(size=14)
    ws["D18"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[18].height=22
    ws.merge_cells("D18:K19")
    return ws
