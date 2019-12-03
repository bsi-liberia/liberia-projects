from maediprojects.lib.spreadsheets import instructions
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import datetime
from io import BytesIO

def guess_types(cell_value):
    if cell_value == None: return ""
    try:
        if float(cell_value) == int(float(cell_value)):
            return int(float(cell_value))
        return float(cell_value)
    except ValueError:
        pass
    try:
        return datetime.datetime.strptime(cell_value, "%Y-%m-%d").date()
    except ValueError:
        pass
    return cell_value

class xlsxDictWriter(object):
    def writesheet(self, worksheet_name):
        """Sheet names can be a maximum of 31 chars. This method
        also writes the header."""
        self.ws = self.wb.create_sheet(worksheet_name[0:30])
        self.writeheader()

    def writerow(self, row_data):
        """Write row, but only for those fields that appear in
        header mapping"""
        hm = self.header_mapping
        for column_header, cell in row_data.items():
            if column_header not in hm: continue
            column_letter = get_column_letter((hm[column_header]))
            self.ws['%s%s'%(column_letter, (self.row_index))] = guess_types(cell)
        self.row_index += 1

    def writeheader(self):
        self.header_mapping.values()
        self.row_index = 1
        self.writerow(dict(map(
            lambda x: (x, x), self.header_mapping.keys()
        )))
        for col in range(1, len(self.header_mapping)+1):
            self.ws['{}1'.format(
                get_column_letter(col)
            )].font = Font(bold=True)

    def delete_first_sheet(self):
        self.wb.remove(self.wb.worksheets[0])
        out = BytesIO()
        self.wb.save(out)
        return out

    def save(self):
        out = BytesIO()
        self.wb.save(out)
        return out

    def write_instructions_sheet(self):
        ws = instructions.cover_sheet(self)
        if self._type == "mtef":
            ws = instructions.mtef(self, ws)
        else:
            ws = instructions.disbursements(self, ws)
        # Protect sheet
        ws.protection.sheet = True

    def __init__(self, headers, _type=u"disbursements",
            template_currency=u"USD",
            instructions_sheet=False):
        self.wb = Workbook()
        self.template_currency = template_currency
        self.instructions_sheet = instructions_sheet
        self._type = _type
        ws = self.wb.worksheets[0]
        ws.title = u"Data"
        if instructions_sheet:
            self.write_instructions_sheet()
        self.header_mapping = dict(
            map(lambda x: (x[1], x[0]), enumerate(headers, start=1)))
