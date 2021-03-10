from openpyxl import load_workbook
import datetime
import sys
from io import BytesIO


def getDataFromFile(f, file_contents, sheetname, by_id=False, headers_row=0):
    headers_row += 1

    if file_contents:
        f = BytesIO(file_contents)
    book = load_workbook(filename=f)

    def getData(book, sheetname, by_id, headers_row):
        if by_id:
            sheet = book[book.sheetnames[sheetname]]
        else:
            sheet = book[sheetname]

        headers = dict(map(lambda cell: (cell[0], cell[1].value), enumerate(sheet[headers_row])))

        def item(row, col):
            return headers[col-1], sheet.cell(row,col).value

        out = [ dict(item(row,col) for col in range(1, sheet.max_column+1)) \
            for row in range(headers_row+1, sheet.max_row+1) ]
        return out
    if sheetname == "all":
        data = []
        for _sheetname in book.sheetnames:
            try:
                data += getData(book, _sheetname, False, headers_row)
            except IndexError:
                break
        return data
    else:
        return getData(book, sheetname, by_id, headers_row)
