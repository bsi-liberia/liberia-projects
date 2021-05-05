from flask import url_for
import pytest
import re
from io import StringIO, BytesIO
from projectdashboard.views.api import JSONEncoder
from projectdashboard.lib.xlsx_to_csv import getDataFromFile
from projectdashboard.query import organisations as qorganisations
import csv
import json
import os
import openpyxl
from werkzeug.datastructures import FileStorage
from projectdashboard.query.generate_xlsx import xlsx_to_csv
from projectdashboard.query import activity as qactivity

@pytest.mark.usefixtures('client_class')
class TestExports:
    def test_nonauth_routes_work(self, client):
        routes = [
            (url_for('exports.activities_csv'), 200),
            (url_for('exports.activities_xlsx_transactions'), 200),
            (url_for('exports.activities_xlsx', domestic_external="external"), 200),
            (url_for('exports.all_activities_xlsx'), 200),
            (url_for('exports.all_activities_xlsx_filtered'), 200),
        ]
        for route, status_code in routes:
            res = client.get(route)
            assert res.status_code == status_code

    def test_auth_routes_work_user(self, client, user, headers_user):
        routes = [
            (url_for('exports.activities_csv'), 200),
            (url_for('exports.activities_xlsx_transactions'), 200),
            (url_for('exports.activities_xlsx', domestic_external="external"), 200),
            (url_for('exports.all_activities_xlsx'), 200),
            (url_for('exports.all_activities_xlsx_filtered'), 200),
        ]
        for route, status_code in routes:
            res = client.get(route, headers=headers_user)
            assert res.status_code == status_code


def write_csv_file(res):
    with open('tests/artefacts/api/exports/activities_csv.csv', 'w') as csvfile:
        remove_last_updated_date(res, csvfile)


def read_csv_file_from_request(res):
    writer = StringIO()
    remove_last_updated_date(res, writer)
    writer.seek(0)
    return writer.read()


def remove_last_updated_date(res, csvfile):
    res_csv_f = StringIO(res.get_data(as_text=True))
    csvreader = csv.DictReader(res_csv_f)
    fieldnames = [fieldname for fieldname in csvreader.fieldnames if fieldname != 'Last updated date']
    csvwriter = csv.DictWriter(csvfile, fieldnames=fieldnames)
    csvwriter.writeheader()
    for row in csvreader:
        csvwriter.writerow(excl_last_updated(row))


def excl_last_updated(row):
    try:
        row.pop('Last updated date')
    except KeyError:
        pass
    return row


def get_clean_data_from_file(xlsx_filename, sheetname='Sheet1', by_id=False):
    if type(xlsx_filename) == BytesIO:
        data = getDataFromFile(xlsx_filename, xlsx_filename.read(), sheetname, by_id)
    else:
        with open(xlsx_filename, 'rb') as xlsxfile:
            data = getDataFromFile(xlsx_filename, xlsxfile.read(), sheetname, by_id)
    return list(map(lambda row: excl_last_updated(row), data))


@pytest.mark.usefixtures('client_class')
class TestExportFiles:
    def test_activities_csv(self, client):
        route = url_for('exports.activities_csv')
        res = client.get(route)

        # NB https://docs.python.org/3/library/csv.html#csv.Dialect.lineterminator
        # write_csv_file(res)

        with open('tests/artefacts/api/exports/activities_csv.csv', 'r') as csvfile:
            in_csv = re.sub("\r\n", "\n", read_csv_file_from_request(res))
            assert in_csv == csvfile.read()

    def test_exports_xlsx(self, client):
        route = url_for('exports.activities_xlsx', domestic_external="external")
        res = client.get(route)
        xlsx_file = BytesIO(res.get_data())
        xlsx_file_as_json = getDataFromFile(xlsx_file, xlsx_file.read(), 'Data1')
        assert len(xlsx_file_as_json) == 10

    def test_exports_xlsx_file(self, client):
        route = url_for('exports.activities_xlsx', domestic_external="external")
        res = client.get(route)
        #with open('tests/artefacts/api/exports/activities_xlsx.xlsx', 'wb') as f:
        #    f.write(res.get_data())
        xlsx_saved = get_clean_data_from_file('tests/artefacts/api/exports/activities_xlsx.xlsx', 'Data1')
        f = BytesIO(res.get_data())
        xlsx_res = get_clean_data_from_file(f, 'Data1')

        assert xlsx_res == xlsx_saved

    def test_export_template(self, client, admin, headers_admin):
        organisation_id = qorganisations.get_organisation_by_name("African Development Bank").id
        route = url_for('exports.export_donor_template',
            reporting_organisation_id=organisation_id,
            currency_code="USD",
            template="disbursements",
            headers="ID,Project code,Activity Title,2019 Q1 (D),Activity Status,Activity Dates (Start Date),Activity Dates (End Date),County")
        res = client.get(route, headers=headers_admin)
        assert res.status_code == 200
        #with open('tests/artefacts/api/exports/export_donor_template.xlsx', 'wb') as f:
        #    f.write(res.get_data())
        xlsx_saved = get_clean_data_from_file('tests/artefacts/api/exports/export_donor_template.xlsx', 1, True)
        f = BytesIO(res.get_data())
        xlsx_res = get_clean_data_from_file(f, 1, True)

        assert xlsx_res == xlsx_saved

    def test_export_template_formatting(self, client, admin, headers_admin):
        organisation_id = qorganisations.get_organisation_by_name("African Development Bank").id
        route = url_for('exports.export_donor_template',
            reporting_organisation_id=organisation_id,
            currency_code="USD",
            template="disbursements",
            headers="ID,Project code,Activity Title,2019 Q1 (D),Activity Status,Activity Dates (Start Date),Activity Dates (End Date),County")
        res = client.get(route, headers=headers_admin)
        assert res.status_code == 200
        f = BytesIO(res.get_data())
        book = openpyxl.load_workbook(f)
        sheet = book[book.sheetnames[1]]

        # Check Activity Status validation
        activity_status_validation = list(filter(lambda validation: str(validation.ranges)=='E2:E12',
            sheet.data_validations.dataValidation))[0]
        assert activity_status_validation.promptTitle == "Activity Status"
        assert activity_status_validation.prompt == "Please select from the list"

        # Check Number validation
        number_validation = list(filter(lambda validation: 'D2:D12' in str(validation.ranges),
            sheet.data_validations.dataValidation))[0]
        assert number_validation.type == 'decimal'

        # Check formatting
        number_fill = sheet.cell(6,4).fill
        assert number_fill.bgColor.value == "00FFFF00"

    def test_zero_spend_activity(self, client, admin, headers_admin):
        # Check there are 10 (external) activities
        route = url_for('exports.activities_xlsx', domestic_external="external")
        res = client.get(route)
        f = BytesIO(res.get_data())
        xlsx_res = get_clean_data_from_file(f, 0, True)
        assert len(xlsx_res) == 10
        # Add an activity with 0 spend
        filename = os.path.join("tests", "artefacts", "testdata_0spend.xlsx")
        with open(filename, "rb") as _file:
            _fakeUpload = FileStorage(
                stream=_file,
                filename="testdata_0spend.xlsx")
            data = xlsx_to_csv.getDataFromFile("testdata.xlsx", _file.read(), 0, True)
            for row in data:
                qactivity.create_activity_for_test(row, admin.id)
        # Check there are now 11 (external) activities
        route = url_for('exports.activities_xlsx', domestic_external="external")
        res = client.get(route)
        f = BytesIO(res.get_data())
        xlsx_res = get_clean_data_from_file(f, 0, True)
        assert len(xlsx_res) == 11
        # Now delete that activity
        route = url_for('activities.activity_delete', activity_id=12)
        res = client.post(route, headers=headers_admin)
        assert res.status_code == 200
