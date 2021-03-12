from openpyxl.utils import get_column_letter, column_index_from_string
from maediprojects.lib.codelists import get_codelists_lookups_by_name
from maediprojects.lib.spreadsheets.validation import v_status, v_id, v_date, v_number
from maediprojects.lib.spreadsheets.formatting import yellowFill, orangeFill

def formatting_validation(writer, num_activities, _headers, counterpart_funding_cols, mtef_cols, disb_cols):
    statuses = get_codelists_lookups_by_name()["ActivityStatus"].keys()
    v_status.formula1='"{}"'.format(u",".join(statuses))

    status_heads = list(map(lambda status_col: _headers.index(status_col)+1,
        [u'Activity Status']))
    id_heads = list(map(lambda id_col: _headers.index(id_col)+1,
        [u'ID']))
    date_heads = list(map(lambda date_col: _headers.index(date_col)+1,
        [u'Activity Dates (Start Date)', u'Activity Dates (End Date)']))

    counterpart_heads = list(map(lambda number_col: _headers.index(number_col)+1,
        counterpart_funding_cols))
    mtef_heads = list(map(lambda number_col: _headers.index(number_col)+1,
        mtef_cols))
    disb_heads = list(map(lambda number_col: _headers.index(number_col)+1,
        disb_cols))

    writer.ws.column_dimensions[u"C"].width = 50
    #writer.ws.column_dimensions[u"D"].width = 35
    for rownum in range(1+1, num_activities+2):
        for col in counterpart_heads:
            writer.ws.cell(row=rownum,column=col).fill = orangeFill
            writer.ws.cell(row=rownum,column=col).number_format = u'"USD "#,##0.00'
        for col in mtef_heads + disb_heads:
            writer.ws.cell(row=rownum,column=col).fill = yellowFill
            writer.ws.cell(row=rownum,column=col).number_format = u'"{} "#,##0.00'.format(writer.template_currency)
    for col in mtef_heads + counterpart_heads + disb_heads:
        writer.ws.column_dimensions[get_column_letter(col)].width = 18
        v_number.add('{}2:{}{}'.format(get_column_letter(col),
            get_column_letter(col), num_activities+2))
    for col in date_heads:
        v_date.add('{}2:{}{}'.format(get_column_letter(col),
            get_column_letter(col), num_activities+2))
        writer.ws.column_dimensions[get_column_letter(col)].width = 20
    for col in id_heads:
        v_id.add('{}2:{}{}'.format(get_column_letter(col),
            get_column_letter(col), num_activities+2))
    for col in status_heads:
        v_status.add('{}2:{}{}'.format(get_column_letter(col),
            get_column_letter(col), num_activities+2))
