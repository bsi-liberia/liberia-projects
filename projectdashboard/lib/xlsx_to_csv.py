from io import BytesIO

from openpyxl import load_workbook


def getDataFromFile(f, file_contents, sheetname, by_id=False, headers_row=0):

    if file_contents:
        f = BytesIO(file_contents)
    book = load_workbook(filename=f, read_only=True)

    def getData(book, sheetname, by_id, headers_row):
        if by_id:
            sheet = book[book.sheetnames[sheetname]]
        else:
            sheet = book[sheetname]

        # FIXME don't convert to list - this can be very slow
        # with massive files that have lots of empty lines
        rows = list(sheet.rows)[headers_row:]
        first_row = [cell.value for cell in rows[0]]
        data = []
        empty_rows = 0
        for row in rows[1:]:
            record = {}
            for key, cell in zip(first_row, row):
                if cell.data_type == 's':
                    record[key] = cell.value.strip()
                else:
                    record[key] = cell.value
            # Ignore empty rows
            # Break after 10 empty rows
            if len(list(filter(lambda cell: cell not in [None, ''], list(record.values())))) == 0:
                empty_rows += 1
                if empty_rows >= 10:
                    break
                continue
            data.append(record)
        return data
    if sheetname == "all":
        data = []
        for _sheetname in book.sheetnames:
            try:
                data += getData(book, _sheetname, False, headers_row)
            except IndexError:
                break
        return data
    else:
        return getData(book, sheetname, by_id, headers_row)
