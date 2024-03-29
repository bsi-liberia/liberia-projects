# -*- coding: UTF-8 -*-

import datetime
import re
import sys
from io import BytesIO

from flask import flash
from openpyxl.styles import Protection
import openpyxl
from six import u as unicode

from projectdashboard import models
from projectdashboard.extensions import db
from projectdashboard.query import activity as qactivity
from projectdashboard.query import finances as qfinances
from projectdashboard.query import counterpart_funding as qcounterpart_funding
from projectdashboard.query import exchangerates as qexchangerates
from projectdashboard.query import organisations as qorganisations
from projectdashboard.query.activity_log import activity_updated
from projectdashboard.lib import xlsx_to_csv, util
from projectdashboard.lib.spreadsheet_headers import headers, fr_headers, headers_transactions
from projectdashboard.lib.spreadsheets.validation import v_status, v_id, v_date, v_number
from projectdashboard.lib.spreadsheets import apply_formatting, helpers, xlsx_writer
from projectdashboard.lib.codelist_helpers import codelists
from projectdashboard.lib.codelists import get_codelists_lookups, get_codelists_lookups_by_name, get_codelists_ids_by_name
from projectdashboard.query.generate_csv import activity_to_json, generate_disb_fys, activity_to_transactions_list


def tidy_amount(amount_value):
    if type(amount_value) == int:
        return float(amount_value)
    if type(amount_value) == float:
        return amount_value
    if amount_value is None:
        return 0.00
    amount_value = amount_value.strip()
    amount_value = re.sub(",", "", amount_value)
    if re.match(r'^-?\d*\.?\d*$', amount_value):  # 2000 or -2000
        return (float(amount_value), "USD")
    elif re.match(r"^(\d*)m (\D*)$", amount_value):  # 20m EUR
        result = re.match(r"^(\d*)m (\D*)$", amount_value).groups()
        return (float(result[0])*1000000, unicode(result[1].upper()))
    elif re.match(r"^(\d*) (\D*)$", amount_value):  # 2000 EUR
        result = re.match(r"^(\d*) (\D*)$", amount_value).groups()
        return (float(result[0]), unicode(result[1].upper()))


def clean_value(amount_value):
    if type(amount_value) == int:
        return float(amount_value)
    if type(amount_value) == float:
        return amount_value
    if amount_value is None:
        return 0
    if amount_value.strip() in ("", "-"):
        return 0
    return float(amount_value.strip())


def process_transaction_classifications(activity):
    activity_classification = activity.classification_data["mtef-sector"]["entries"][0]
    classifications = []
    cl = models.ActivityFinancesCodelistCode()
    cl.codelist_id = 'mtef-sector'
    cl.codelist_code_id = activity_classification.codelist_code_id
    classifications.append(cl)
    return classifications


def process_transaction(activity, amount, currency, column_name):
    provider = activity.funding_organisations[0].id
    receiver = activity.implementing_organisations[0].id
    fq, fy = util.get_data_from_header(column_name)
    end_fq_date = util.fq_fy_to_date(fq, fy, "end")
    disbursement = models.ActivityFinances()
    disbursement.transaction_date = end_fq_date
    disbursement.transaction_type = "D"
    disbursement.transaction_description = "Disbursement for Q{} FY{}, imported from AMCU template".format(
        fq, fy
    )
    disbursement.currency = currency
    disbursement.currency_automatic = True
    disbursement.currency_source, disbursement.currency_rate, disbursement.currency_value_date = qexchangerates.get_exchange_rate(
        disbursement.transaction_date, disbursement.currency)
    disbursement.transaction_value_original = amount
    disbursement.provider_org_id = provider
    disbursement.receiver_org_id = receiver
    disbursement.finance_type = activity.finance_type
    disbursement.aid_type = activity.aid_type
    disbursement.classifications = process_transaction_classifications(
        activity)
    return disbursement


friendly_to_column = {
    'Activity Title': 'title',
    'Activity Description': 'description',
    'Activity Dates (Start Date)': 'start_date',
    'Activity Dates (End Date)': 'end_date',
    'SDG Goal': 'sdg-goals'
}


def update_activity_attrib(activity, existing_activity, column, value, codelists, codelists_with_ids):
    print(f"Attempting update of {column}")
    friendly_column = friendly_to_column.get(column)
    if column in ['Activity Title', 'Activity Description']:
        if existing_activity.get(column) != value:
            setattr(activity, friendly_column, value)
            return True
    elif column in ['Implemented by']:
        if existing_activity.get(column) != value:
            for organisation in activity.organisations:
                if organisation.role == 4:
                    db.session.delete(organisation)
            activity.organisations.append(qorganisations.make_organisation(
                value), 4)
            return True
    elif column in ['Donor project code']:
        if (existing_activity.get(column) != str(value)) and (str(value) != ""):
            activity.iati_identifier = value
            return True
    elif column in ['Activity Status']:
        if existing_activity.get(column) != value:
            activity.activity_status = codelists["ActivityStatus"][value]
            return True
    elif column in ['Activity Dates (Start Date)', 'Activity Dates (End Date)']:
        if existing_activity.get(column) != value.date().isoformat():
            setattr(activity, friendly_column, value.date())
        return True
    elif column in ['SDG Goal']:
        if value not in (None, ''):
            sdg_goal = value.split(" - ")[1]
            activity.set_classification('sdg-goals', codelists_with_ids['sdg-goals'][sdg_goal])
            return True


def update_activity_data(activity_cols, activity, existing_activity, row, codelists, codelists_with_ids):
    updated = False
    for column in activity_cols:
        if column not in row:
            print(f"WARNING: column {column} not in row")
        else:
            _updated = update_activity_attrib(activity, existing_activity, column,
                row.get(column), codelists, codelists_with_ids)
            if _updated == True: updated = True
    activity.forwardspends += qfinances.create_missing_forward_spends(
        activity.start_date, activity.end_date, activity.id)
    return updated


def parse_mtef_cols(currency, mtef_cols, existing_activity, row, activity_id):
    updated_years = []
    for mtef_year in mtef_cols:
        new_fy_value = clean_value(row[mtef_year])
        if 'Q' in mtef_year:  # Quarterly MTEF projections
            _mtef_year_year, _mtef_year_quarter = re.match(
                r"FY(\S*) Q(\d*) \(MTEF\)", mtef_year).groups()
            existing_fy_value = float(existing_activity["FY{} Q{} (MTEF)".format(
                _mtef_year_year, _mtef_year_quarter)])

            new_fy_value_in_usd = qexchangerates.convert_from_currency(
                currency=currency,
                _date=datetime.datetime.utcnow().date(),
                value=new_fy_value)
            difference = new_fy_value_in_usd-existing_fy_value
            # We ignore differences < 1 USD, because this can be due to rounding errors
            # when we divided input data by 4.
            if round(difference) == 0:
                continue
            value = new_fy_value_in_usd
            year, quarter = util.lr_quarter_to_cal_quarter(
                int("{}".format(_mtef_year_year)), int(_mtef_year_quarter))
            inserted = qfinances.create_or_update_forwardspend(
                activity_id, quarter, year, value, "USD")
            updated_years.append("FY{} Q{}".format(
                _mtef_year_year, _mtef_year_quarter))
        else:  # Annual MTEF projections

            fy_start = re.match(r"FY(\S*) \(MTEF\)", mtef_year).groups()[0]
            fy = models.FiscalYear.query.filter_by(
                id='FY{}'.format(fy_start)).first()

            existing_fy_value = sum([float(existing_activity["FY{} Q{} (MTEF)".format(
                fy_start, quarter)]) for quarter in range(1, fy.num_quarters+1)])

            new_fy_value_in_usd = qexchangerates.convert_from_currency(
                currency=currency,
                _date=datetime.datetime.utcnow().date(),
                value=new_fy_value)
            difference = new_fy_value_in_usd-existing_fy_value
            # We ignore differences < 1 USD, because this can be due to rounding errors
            # when we divided input data by 4.
            if round(difference) == 0:
                continue
            # Create 1/4 of new_fy_value for each quarter in this FY
            value = round(new_fy_value_in_usd/float(fy.num_quarters), 4)
            for fp in fy.fiscal_year_periods:
                year, quarter = util.date_to_fy_fq_calendar_year(fp.start)
                inserted = qfinances.create_or_update_forwardspend(
                    activity_id, quarter, year, value, "USD")
            updated_years.append("FY{}".format(fy_start))
    return updated_years


def parse_counterpart_cols(counterpart_funding_cols, activity, row, activity_id):
    updated_counterpart_years = []
    for counterpart_year in counterpart_funding_cols:
        new_fy_value = clean_value(row[counterpart_year])
        cfy_start = re.match(
            r"FY(\S*) \(GoL counterpart fund request\)", counterpart_year).groups()[0]
        fy = models.FiscalYear.query.filter_by(
            id='FY{}'.format(cfy_start)).first()
        existing_cfy_value = activity.FY_counterpart_funding_for_FY(
            "FY{}".format(cfy_start))
        difference = new_fy_value-existing_cfy_value
        if difference == 0:
            continue
        inserted = qcounterpart_funding.create_or_update_counterpart_funding(
            activity_id, fy.start, new_fy_value)
        updated_counterpart_years.append("FY{}".format(cfy_start))
    return updated_counterpart_years


def parse_disbursement_cols(currency, disbursement_cols, activity, existing_activity, row):
    updated_disbursements = []
    for column_name in disbursement_cols:
        row_value = clean_value(row[column_name])
        fq, fy = util.get_data_from_header(column_name)
        column_date = util.fq_fy_to_date(fq, fy, "end")
        existing_value = float(existing_activity.get(column_name, 0))
        existing_value_same_currency = qexchangerates.convert_to_currency(
            currency=currency,
            _date=column_date,
            value=existing_value)
        difference = round(row_value - existing_value_same_currency, 4)
        if (round(difference) == 0):
            continue
        activity.finances.append(
            process_transaction(activity, difference, currency, column_name)
        )
        db.session.add(activity)
        if existing_activity.get(column_name, 0) != 0:
            # Non-zero financial values were previously provided
            # and should be adjusted upwards/downwards
            updated_disbursements.append("{}; previous value was {}; \
                new value is {}; new entry for {} added".format(
                util.column_data_to_string(column_name),
                existing_value_same_currency,
                row_value,
                difference))
        else:
            # Financial values were not previously provided, and are now entered
            updated_disbursements.append("{}".format(
                util.column_data_to_string(column_name)))
    return updated_disbursements


def make_updated_info(updated, activity, num_updated_activities):
    if (updated.get("mtef_years") or updated.get("counterpart_years") or
        updated.get("disbursements") or updated.get("activity")):
        num_updated_activities += 1
        activity_updated(activity.id)
    else:
        return None, num_updated_activities
    msg = "Updated {} (Project ID: {}): ".format(
        activity.title,
        activity.id)
    msgs = []
    if updated.get("activity"):
        msgs.append("updated activity data")
    if updated.get("mtef_years"):
        msgs.append("updated MTEF projections for {}".format(
            ", ".join(updated["mtef_years"])))
    if updated.get("counterpart_years"):
        msgs.append("updated counterpart funding for {}".format(
            ", ".join(updated["counterpart_years"])))
    if updated.get("disbursements"):
        msgs.append("updated disbursement data for {}".format(
            ", ".join(updated["disbursements"])))
    return msg + "; ".join(msgs), num_updated_activities


def import_xls_mtef(input_file, column_names, activity_column_names):
    return import_xls_new(input_file, "mtef", column_names, activity_column_names)


def import_xls(input_file, column_names, activity_column_names):
    return import_xls_new(input_file, "disbursements", column_names, activity_column_names)


def import_xls_new(input_file, _type, disbursement_cols=[], activity_cols=[]):
    num_updated_activities = 0
    messages = []
    activity_id = None
    file_contents = BytesIO(input_file.read())
    xl_workbook = openpyxl.load_workbook(file_contents)
    num_sheets = len(xl_workbook.sheetnames)
    cl_lookups = get_codelists_lookups()
    cl_lookups_by_name = get_codelists_lookups_by_name()
    cl_lookups_by_name_id = get_codelists_ids_by_name()

    def filter_mtef(column):
        pattern = r"(\S*) \(MTEF\)$"
        return re.match(pattern, column)

    def filter_counterpart(column):
        pattern = r"(\S*) \(GoL counterpart fund request\)$"
        return re.match(pattern, column)
    if "Instructions" in xl_workbook.sheetnames:
        currency = xl_workbook["Instructions"].cell(6, 3).value
        begin_sheet = 1
    else:
        currency = "USD"
        begin_sheet = 0
    try:
        for sheet_id in range(begin_sheet, num_sheets):
            input_file.seek(0)
            data = xlsx_to_csv.getDataFromFile(
                input_file.filename, input_file.read(), sheet_id, True)
            if _type == 'mtef':
                mtef_cols = list(filter(filter_mtef, data[0].keys()))
                counterpart_funding_cols = list(
                    filter(filter_counterpart, data[0].keys()))
                if len(mtef_cols) == 0:
                    raise Exception("No columns containing MTEF projections data \
                    were found in the uploaded spreadsheet!")
            elif _type == 'disbursements':
                for _column_name in disbursement_cols:
                    if _column_name not in data[0].keys():
                        raise Exception("The column {} containing financial data was not \
                        found in the uploaded spreadsheet!".format(_column_name))
            for row in data:  # each row is one ID
                try:
                    activity_id = int(row["ID"])
                except TypeError:
                    messages.append("Warning, activity ID \"{}\" with title \"{}\" was not found in the system \
                        and was not imported! Please create this activity in the \
                        system before trying to import.".format(row['ID'], row['Activity Title']))
                    continue
                activity = qactivity.get_activity(activity_id)
                if activity is None:
                    messages.append("Warning, activity ID \"{}\" with title \"{}\" was not found in the system \
                        and was not imported! Please create this activity in the \
                        system before trying to import.".format(row['ID'], row['Activity Title']))
                    continue
                activity_iati_preferences = [
                    pref.field for pref in activity.iati_preferences]
                if not activity:
                    messages.append("Warning, activity ID \"{}\" with title \"{}\" was not found in the system \
                        and was not imported! Please create this activity in the \
                        system before trying to import.".format(row['ID'], row['Activity Title']))
                    continue
                existing_activity = activity_to_json(activity, cl_lookups)
                if _type == 'mtef':
                    # FIXME quick fix for now
                    if 'forwardspend' in activity_iati_preferences:
                        updated = {
                            'activity': update_activity_data(activity_cols, activity, existing_activity,
                                row, cl_lookups_by_name, cl_lookups_by_name_id)
                        }
                        continue
                    updated = {
                        'activity': update_activity_data(activity_cols, activity, existing_activity,
                            row, cl_lookups_by_name, cl_lookups_by_name_id),
                        # Parse MTEF projections columns
                        'mtef_years': parse_mtef_cols(currency, mtef_cols, existing_activity,
                            row, activity_id),
                        # Parse counterpart funding columns
                        'counterpart_years': parse_counterpart_cols(counterpart_funding_cols,
                            activity, row, activity_id),
                    }
                elif _type == 'disbursements':
                    # FIXME quick fix for now
                    if 'disbursement' in activity_iati_preferences:
                        updated = {
                            'activity': update_activity_data(activity_cols, activity, existing_activity,
                                row, cl_lookups_by_name, cl_lookups_by_name_id)
                        }
                        continue
                    updated = {
                        'activity': update_activity_data(activity_cols, activity, existing_activity,
                            row, cl_lookups_by_name, cl_lookups_by_name_id),
                        'disbursements': parse_disbursement_cols(currency, disbursement_cols,
                            activity, existing_activity, row)
                    }
                # Mark activity as updated and inform user
                update_message, num_updated_activities = make_updated_info(
                    updated, activity, num_updated_activities)
                if update_message is not None:
                    messages.append(update_message)
    except Exception as e:
        if activity_id is not None:
            messages.append("""There was an unexpected error when importing your
            projects, there appears to be an error around activity ID {}.
            The error was: {}""".format(activity_id, e))
        else:
            messages.append("""There was an error while importing your projects,
        the error was: {}""".format(e))
    db.session.commit()
    return messages, num_updated_activities


def generate_xlsx_filtered(arguments={}):
    disbFYs = generate_disb_fys()
    _headers = headers + disbFYs
    writer = xlsx_writer.xlsxDictWriter(_headers)
    writer.writesheet("Data")
    writer.writeheader()
    cl_lookups = get_codelists_lookups()
    activities = qactivity.list_activities_by_filters(
        arguments)
    for activity in activities:
        total_commitments = activity.total_commitments
        total_disbursements = activity.total_disbursements
        for fundsource in activity.fund_sources:
            fund_source_code = fundsource.code
            fund_source_name = fundsource.name
            fund_source_commitments = sum(map(
                lambda item: item['value'], activity.FY_commitments_dict_fund_sources[fundsource.id].values()))
            fund_source_disbursements = sum(map(
                lambda item: item['value'], activity.FY_disbursements_dict_fund_sources[fundsource.id].values()))
            # MTEF projections don't have fund sources. So we pro-rate their value out, ideally
            # by commitments, alternatively by disbursements.
            # This is a bit of an ugly hack and this whole function needs to be rewritten
            # in the near future.
            if (total_commitments is not None) and (total_commitments > 0):
                pct = fund_source_commitments/total_commitments
            elif (total_disbursements is not None) and (total_disbursements > 0):
                pct = fund_source_disbursements/total_disbursements
            else:
                pct = 1
            activity_data = activity_to_json(activity, cl_lookups)
            activity_data.update(
                dict(map(lambda d: (d, 0.00), list(generate_disb_fys()))))
            activity_data["Fund Source"] = "{} - {}".format(fund_source_code,
                fund_source_name) if fund_source_name != fund_source_code else fund_source_name
            # Add Disbursements data
            if fundsource.id is not None and fundsource.finance_type:
                activity_data['Finance Type (Type of Assistance)'] = {'110': 'Grant', '410': 'Loan'}[fundsource.finance_type]
                activity_data['Total Commitments'] = activity_data['Total Commitments'] * pct
                activity_data['Total Disbursements'] = activity_data['Total Disbursements'] * pct
            if fundsource.id in activity.FY_commitments_dict_fund_sources:
                activity_data.update(dict(map(lambda d: (
                    d[0], d[1]["value"]), activity.FY_commitments_dict_fund_sources[fundsource.id].items())))
            if fundsource.id in activity.FY_disbursements_dict_fund_sources:
                activity_data.update(dict(map(lambda d: (
                    d[0], d[1]["value"]), activity.FY_disbursements_dict_fund_sources[fundsource.id].items())))
            activity_data.update(dict(map(lambda d: (
                d[0], d[1]["value"]*pct), activity.FY_forward_spend_dict_fund_sources[None].items())))
            writer.writerow(activity_data)
    writer.delete_first_sheet()
    return writer.save()


def generate_xlsx_export_template(data, mtef=False, currency="USD", _headers=None):
    mtef_cols, counterpart_funding_cols, disb_cols, _headers = helpers.get_column_information(
        mtef, _headers)
    for required_field in ["ID", "Activity Status", 'Activity Dates (Start Date)', 'Activity Dates (End Date)']:
        if required_field not in _headers:
            flash("Error: the field `{}` is required in this export. Please adjust your selected fields and try again!".format(
                required_field))
            return False
    writer = xlsx_writer.xlsxDictWriter(_headers,
                                        _type={True: "mtef",
                                               False: "disbursements"}[mtef],
                                        template_currency=currency,
                                        instructions_sheet=True)
    cl_lookups = get_codelists_lookups()

    for org_code, activities in sorted(data.items()):
        writer.writesheet(org_code)
        writer.ws.add_data_validation(v_status)
        writer.ws.add_data_validation(v_date)
        writer.ws.add_data_validation(v_number)
        writer.ws.add_data_validation(v_id)
        #writer.ws.protection.sheet = True
        for activity in activities:
            existing_activity = activity_to_json(activity, cl_lookups)
            for mtef_year in mtef_cols:
                fy_start = re.match(r"FY(\S*) \(MTEF\)", mtef_year).groups()[0]
                # Every FY has at least one quarter, but may not have more than one quarter.
                existing_activity[mtef_year] = sum([float(existing_activity["FY{} Q1 (MTEF)".format(fy_start)]),
                                                    float(existing_activity.get(
                                                        "FY{} Q2 (MTEF)".format(fy_start), 0)),
                                                    float(existing_activity.get(
                                                        "FY{} Q3 (MTEF)".format(fy_start), 0)),
                                                    float(existing_activity.get("FY{} Q4 (MTEF)".format(fy_start), 0))])
                if writer.template_currency != "USD":
                    # N.B.: we convert at (close to) today's date,
                    # because these are *projections* and we store in USD
                    existing_activity[mtef_year] = qexchangerates.convert_to_currency(
                        currency=writer.template_currency,
                        _date=datetime.datetime.utcnow().date(),
                        value=existing_activity[mtef_year])
            # Convert disbursement data to writer.template_currency
            for disb_year in disb_cols:
                if writer.template_currency != "USD":
                    existing_activity[disb_year] = qexchangerates.convert_to_currency(
                        currency=writer.template_currency,
                        _date=util.get_real_date_from_header(disb_year, "end"),
                        value=existing_activity[disb_year])
            # Leave in USD always
            for counterpart_year in counterpart_funding_cols:
                cfy = re.match(
                    r"FY(\S*) \(GoL counterpart fund request\)", counterpart_year).groups()[0]
                existing_activity[counterpart_year] = activity.FY_counterpart_funding_for_FY(
                    "FY{}".format(cfy))
            writer.writerow(existing_activity)
        # Formatting
        apply_formatting.formatting_validation(writer, len(
            activities), _headers, counterpart_funding_cols, mtef_cols, disb_cols)
    writer.delete_first_sheet()
    return writer.save()


def generate_xlsx_transactions(filter_key=None, filter_value=None):
    disbFYs = generate_disb_fys()
    writer = xlsx_writer.xlsxDictWriter(headers_transactions)
    writer.writesheet("Data")
    writer.writeheader()
    cl_lookups = get_codelists_lookups()
    if (filter_key and filter_value):
        activities = qactivity.list_activities_by_filters(
            {filter_key: filter_value})
    else:
        activities = qactivity.list_activities()
    for activity in activities:
        for tr in activity_to_transactions_list(activity, cl_lookups):
            writer.writerow(tr)
    writer.delete_first_sheet()
    return writer.save()
